from openpyxl import load_workbook
from utils.printUtils import printWithTs
import datetime as dt
import pandas as pd
import openpyxl


def pasteScadaAvailToTemplateFile(reqDt: dt.datetime, templateFile: str, shtName: str):
    printWithTs('loading data file data into a dataframe', clr='magenta')
    # read input data from file in the format input_data/rtu_avail_reports/RTUs Availability Reports_29_01_2025.xlsx
    fileDateStr: str = reqDt.strftime("%d_%m_%Y")
    inpFilePath: str = f"input_data/rtu_avail_reports/RTUs Availability Reports_{fileDateStr}.xlsx"
    inpDf = pd.read_excel(inpFilePath)

    # load report template workbook sheet 'Daily REMC Report_Part2'
    printWithTs('loading template file', clr='magenta')
    templWb = load_workbook(templateFile)
    dataSht = templWb[shtName]
    # search the starting row address by the named cell "scada_avail_start"
    reportStartCoord = list(templWb.defined_names["scada_avail_start"]
                            .destinations)[0][1].replace("$", "")
    (startInd, colInd) = openpyxl.utils.cell.coordinate_to_tuple(reportStartCoord)

    # Paste the data from input data in the template rows in the sequence (name, availability percentage, duration=24hrs*0.01*avail_perc)
    printWithTs('started pasting data to template', clr='magenta')
    for dfRowInd in range(len(inpDf)):
        dataSht.cell(startInd+dfRowInd, colInd).value = inpDf.iloc[dfRowInd, 0]
        availPerc = inpDf.iloc[dfRowInd, 1]
        availHrs = availPerc*24*0.01
        dataSht.cell(startInd+dfRowInd, colInd + 1).value = availPerc
        dataSht.cell(startInd+dfRowInd, colInd + 2).value = formatHrs(availHrs)
    printWithTs('Saving report template after pasting')
    templWb.save(str(templateFile))


def formatHrs(hrs: float) -> str:
    hrsInt: int = int(hrs)
    hrsStr: str = str(hrsInt)
    hrsStr = "0"+hrsStr if hrsInt < 10 else hrsStr
    minsInt: int = int((hrs % 1)*60)
    minsStr: str = str(minsInt)
    minsStr = "0"+minsStr if minsInt < 10 else minsStr
    return f"{hrsStr}:{minsStr}"
