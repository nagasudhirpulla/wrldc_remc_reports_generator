from openpyxl import load_workbook
from utils.printUtils import printWithTs
from utils.dateUtils import getReportForDate
import datetime as dt


def pasteDataToTemplateFile(dataFile, templateFile):
    printWithTs('loading data file data into a dataframe', clr='magenta')
    # open an Excel file and return a workbook
    dataWb = load_workbook(dataFile, read_only=True)
    dataFileSheetNames = [sht for sht in dataWb.sheetnames]

    printWithTs('loading template file', clr='magenta')
    templWb = load_workbook(templateFile)
    templFileSheetNames = [sht for sht in templWb.sheetnames]

    printWithTs('started pasting data to template', clr='magenta')
    # print(dataFileSheetNames)
    # iterate through each sheet of data file for copying
    # https://www.geeksforgeeks.org/python-how-to-copy-data-from-one-excel-sheet-to-another/
    for dataShtName in dataFileSheetNames:
        printWithTs('pasting data sheet {0}'.format(dataShtName))
        # skip sheet of data file if same sheet absent in template file
        if not(dataShtName in templFileSheetNames):
            printWithTs('skipping sheet paste as not present in template file')
            continue

        # copy paste cell values to template sheet
        dataSht = dataWb[dataShtName]
        maxRows = dataSht.max_row
        maxCols = dataSht.max_column
        printWithTs('maxRows = {0}, maxCols = {1}'.format(maxRows, maxCols))

        for rowIter in range(1, maxRows+1):
            # check if 1st column value is 'dummy**'
            if dataSht.cell(row=rowIter, column=1).value == 'dummy**':
                continue
            # copy cell values from data sheet to template sheet
            for colIter in range(1, maxCols+1):
                printWithTs('pasting row={0}, col={1}'.format(rowIter, colIter))
                cellVal = dataSht.cell(row=rowIter, column=colIter).value
                templWb[dataShtName].cell(
                    row=rowIter, column=colIter).value = cellVal

    printWithTs('done pasting data to template', clr='green')
    # TODO avoid hard coding
    templWb["Daily REMC Report_Part1"]["K3"] = getReportForDate()+dt.timedelta(days=1)
    # saving the destination excel file
    printWithTs('Saving report template after pasting')
    templWb.save(str(templateFile))
    dataWb.close()
