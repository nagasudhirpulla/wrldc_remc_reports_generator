from openpyxl import load_workbook
import pandas as pd
from openpyxl import Workbook
import os

## todo create excel if not exists
def createExcelIfAbsent(filePath):
    if not(os.path.isfile(filePath)):
        # create workbook
        wb = Workbook()
        wb.save(filename = filePath)

def deleteSheetIfExists(filePath, sheetName):
    wb = load_workbook(filePath)   # open an Excel file and return a workbook
    if sheetName in wb.sheetnames:
        # delete sheet
        wb.remove_sheet(wb.get_sheet_by_name(sheetName))
        wb.save(filePath)
    wb.close()

def saveDfToExcelSheet(filePath, sheetName, dataDf, deleteSheet=False):
    createExcelIfAbsent(filePath)
    if deleteSheet == True:
        deleteSheetIfExists(filePath, sheetName)
    with pd.ExcelWriter(filePath, mode='a') as writer: # pylint: disable=abstract-class-instantiated
        dataDf.to_excel(writer, sheet_name=sheetName, index=False)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    updated function available at https://stackoverflow.com/questions/66531396/export-pandas-dataframe-to-xlsx-dealing-with-the-openpyxl-issue-on-python-3-9
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # edit by Nagasudhir as per requirement
    if startrow == None and truncate_sheet == True:
        startrow = 0
    # edit by Nagasudhir as per requirement
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') # pylint: disable=abstract-class-instantiated

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()